import obspython as obs
import csv
import io
import os
import re
import subprocess
import sys
import threading
import time
import urllib.request

try:
    import openpyxl
    deps_missing = []
except ImportError:
    deps_missing = ["openpyxl"]

excel_file = ""
worksheet = ""
source_type_column = ""
source_name_column = ""
value_column = ""
db_format = "excel"
gsheet_url = ""
gsheet_creds_path = ""
is_running = False
reload_interval = 3
last_file_modified = 0
data_cache = None
update_thread = None
stop_event = threading.Event()
status_text = "Stopped"
install_requested = set()

KEY_FOR_TYPE = {"text": "text", "image": "file", "media": "local_file", "browser": "url"}
TRUTHY = ("true", "1", "yes", "show", "on")
COL_KEYS = ("source_type_column", "source_name_column", "value_column")


def script_description():
    """Returns the description shown in OBS"""
    return (
        "<b>Enhanced Excel/CSV Source Controller</b>"
        "<hr>"
        "Automatically updates OBS sources based on Excel/CSV/Google Sheet content."
        "<br><br>"
        "Columns needed:"
        "<ul>"
        "<li>Source Type (text/image/media/browser/color/visibility/scene)</li>"
        "<li>Source Name (as named in OBS; scene name for 'scene' rows)</li>"
        "<li>Value (text/path/URL, #RRGGBB for color, TRUE/FALSE for visibility and scene)</li>"
        "</ul>"
        "By John Paolo 'CHOCO!' Baclayon <br>"
        "This program is free to use and free to modify under the MIT License"
    )


def find_python():
    if sys.platform == "win32":
        return os.path.join(sys.exec_prefix, "python.exe")
    return os.path.join(sys.exec_prefix, "bin", "python3")


def install_deps(packages):
    global status_text
    packages = [p for p in packages if p not in install_requested]
    if not packages:
        return
    install_requested.update(packages)

    def run():
        global status_text
        status_text = f"Installing {' '.join(packages)}..."
        print(status_text)
        try:
            subprocess.check_call([find_python(), "-m", "pip", "install"] + packages)
            status_text = "Dependencies installed — reload the script (Scripts > Reload)"
        except Exception as e:
            status_text = f"Install failed ({e}). Run manually: pip install {' '.join(packages)}"
        print(status_text)

    threading.Thread(target=run, daemon=True).start()


def gsheet_export_url(url):
    m = re.search(r"/d/([\w-]+)", url or "")
    if not m:
        return None
    gid = re.search(r"gid=(\d+)", url)
    return (f"https://docs.google.com/spreadsheets/d/{m.group(1)}"
            f"/export?format=csv&gid={gid.group(1) if gid else 0}")


def parse_color(value):
    s = str(value).strip().lstrip("#")
    r, g, b = int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
    return 0xFF000000 | (b << 16) | (g << 8) | r


def parse_csv_text(text):
    reader = csv.reader(io.StringIO(text))
    headers = [h.strip() for h in next(reader, [])]
    return headers, [dict(zip(headers, row)) for row in reader]


def load_gsheet_api():
    global status_text
    try:
        import gspread
    except ImportError:
        install_deps(["gspread", "google-auth"])
        status_text = "Installing gspread for the Sheets API — reload the script when done"
        return None
    gc = gspread.service_account(filename=gsheet_creds_path)
    sh = gc.open_by_url(gsheet_url)
    gid = re.search(r"gid=(\d+)", gsheet_url)
    ws = sh.get_worksheet_by_id(int(gid.group(1))) if gid else sh.sheet1
    values = ws.get_all_values()
    headers = [h.strip() for h in values[0]] if values else []
    return headers, [dict(zip(headers, row)) for row in values[1:]]


def load_data():
    """Returns (headers, rows-as-dicts) or None on failure (caller keeps cache and retries)."""
    global status_text
    try:
        if db_format == "gsheet":
            if gsheet_creds_path:
                return load_gsheet_api()
            export = gsheet_export_url(gsheet_url)
            if not export:
                status_text = "Invalid Google Sheet URL"
                return None
            with urllib.request.urlopen(export, timeout=10) as resp:
                return parse_csv_text(resp.read().decode("utf-8-sig"))

        if not os.path.isfile(excel_file):
            status_text = f"File not found: {excel_file or '(no file set)'}"
            return None
        # ponytail: one quick byte read keeps the lock window at ms so Excel can
        # save while we run; if Excel still reports 'in use', the upgrade path is
        # CreateFileW with FILE_SHARE_DELETE via ctypes.
        with open(excel_file, "rb") as f:
            raw = f.read()
        if db_format == "csv":
            return parse_csv_text(raw.decode("utf-8-sig"))
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[worksheet] if worksheet in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
        rows = [dict(zip(headers, row)) for row in rows_iter]
        wb.close()
        return headers, rows
    except (PermissionError, OSError) as e:
        status_text = f"File busy/unreachable, retrying: {e}"
        print(status_text)
        return None
    except Exception as e:
        status_text = f"Failed to load data: {e}"
        print(status_text)
        return None


def update_obs_sources():
    global last_file_modified, data_cache, status_text

    if not is_running:
        return

    if db_format in ("excel", "csv") and os.path.isfile(excel_file):
        try:
            current_modified = os.path.getmtime(excel_file)
        except OSError:
            current_modified = None
        if current_modified == last_file_modified and data_cache is not None:
            data = data_cache
        else:
            data = load_data()
            if data is None:
                data = data_cache  # keep last good data; failed read retries next interval
            else:
                data_cache = data
                last_file_modified = current_modified
    else:
        data = load_data()  # ponytail: gsheet refetches every interval; add content-hash skip if it ever matters
        if data is not None:
            data_cache = data
        else:
            data = data_cache

    if data is None:
        return  # status_text already explains why

    headers, rows = data
    for col in (source_type_column, source_name_column, value_column):
        if col not in headers:
            status_text = f"Column '{col}' not found — check column settings"
            print(status_text)
            return

    updated = skipped = errors = 0
    for i, row in enumerate(rows):
        try:
            source_type = row.get(source_type_column)
            source_name = row.get(source_name_column)
            value = row.get(value_column)
            # ponytail: blank cells never touch OBS — kills the endless failures
            # from template rows and 'nan' text on stream
            if any(v is None or str(v).strip() == "" for v in (source_type, source_name, value)):
                skipped += 1
                continue
            type_lower = str(source_type).strip().lower()
            source_name = str(source_name).strip()
            if type_lower in KEY_FOR_TYPE:
                update_source_setting(source_name, KEY_FOR_TYPE[type_lower], str(value))
                updated += 1
            elif type_lower == "color":
                update_source_setting(source_name, "color", parse_color(value), as_int=True)
                updated += 1
            elif type_lower == "visibility":
                set_source_visibility(source_name, str(value).strip().lower() in TRUTHY)
                updated += 1
            elif type_lower == "scene":
                if str(value).strip().lower() in TRUTHY and switch_scene(source_name):
                    updated += 1
            else:
                skipped += 1
        except Exception as e:
            print(f"Row {i + 2}: Error - {e}")
            errors += 1
    status_text = f"{time.strftime('%H:%M:%S')} — updated {updated}, skipped {skipped}, errors {errors}"
    if errors:
        print(status_text)


def threaded_updater():
    update_obs_sources()
    while not stop_event.wait(reload_interval):
        update_obs_sources()


def update_source_setting(source_name, key, value, as_int=False):
    source = obs.obs_get_source_by_name(source_name)
    if not source:
        return
    settings = obs.obs_data_create()
    if as_int:
        obs.obs_data_set_int(settings, key, value)
    else:
        obs.obs_data_set_string(settings, key, value)
    obs.obs_source_update(source, settings)
    obs.obs_data_release(settings)
    obs.obs_source_release(source)


def set_source_visibility(source_name, visible):
    # ponytail: toggles the item in every scene containing it; add a scene
    # column if per-scene control is ever needed
    scenes = obs.obs_frontend_get_scenes()
    for scene_source in scenes:
        scene = obs.obs_scene_from_source(scene_source)
        if scene:
            item = obs.obs_scene_find_source_recursive(scene, source_name)
            if item:
                obs.obs_sceneitem_set_visible(item, visible)
    obs.source_list_release(scenes)


def switch_scene(scene_name):
    current = obs.obs_frontend_get_current_scene()
    current_name = obs.obs_source_get_name(current)
    obs.obs_source_release(current)
    if current_name == scene_name:
        return False  # already live; don't re-trigger the transition every interval
    target = obs.obs_get_source_by_name(scene_name)
    if not target:
        return False
    obs.obs_frontend_set_current_scene(target)
    obs.obs_source_release(target)
    return True


def probe_file(path, fmt, sheet):
    """Sheet names and column headers for the picker dropdowns; ([], []) if unreadable."""
    try:
        if fmt == "csv" and os.path.isfile(path):
            with open(path, "rb") as f:
                first_line = f.readline().decode("utf-8-sig")
            return ["CSV"], [h.strip() for h in next(csv.reader(io.StringIO(first_line)), [])]
        if fmt == "excel" and os.path.isfile(path):
            with open(path, "rb") as f:
                raw = f.read()
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active
            headers = [str(h).strip() for h in next(ws.iter_rows(values_only=True), []) if h is not None]
            names = wb.sheetnames
            wb.close()
            return names, headers
    except Exception as e:
        print(f"Could not read sheet/column names: {e}")
    return [], []


def fill_list(prop, options, current):
    obs.obs_property_list_clear(prop)
    for o in options:
        obs.obs_property_list_add_string(prop, o, o)
    if current and current not in options:
        obs.obs_property_list_add_string(prop, current, current)  # never wipe a stored setting


def refresh_pickers(props, prop, settings):
    """Repopulates the sheet/column dropdowns from the selected file (gsheet: type manually)."""
    if settings is not None:
        path = obs.obs_data_get_string(settings, "excel_file")
        fmt = obs.obs_data_get_string(settings, "db_format")
        sheet = obs.obs_data_get_string(settings, "worksheet")
        cols = {k: obs.obs_data_get_string(settings, k) for k in COL_KEYS}
    else:  # initial fill when the dialog opens, from the loaded settings
        path, fmt, sheet = excel_file, db_format, worksheet
        cols = dict(zip(COL_KEYS, (source_type_column, source_name_column, value_column)))
    sheets, headers = probe_file(path, fmt, sheet)
    fill_list(obs.obs_properties_get(props, "worksheet"), sheets, sheet)
    for k in COL_KEYS:
        fill_list(obs.obs_properties_get(props, k), headers, cols[k])
    return True


def refresh_status(props, prop):
    obs.obs_property_set_description(obs.obs_properties_get(props, "status_info"), status_text)
    return True


def toggle_script(props, prop):
    global is_running, update_thread, status_text
    if not is_running:
        if deps_missing:
            status_text = "Waiting on dependency install — reload the script once it finishes"
        else:
            is_running = True
            stop_event.clear()
            update_thread = threading.Thread(target=threaded_updater, daemon=True)
            update_thread.start()
            status_text = "Running"
    else:
        is_running = False
        stop_event.set()  # ponytail: no join — daemon thread exits on its own; joining freezes the OBS UI
        status_text = "Stopped"
    print(f"Script is now: {status_text}")
    return refresh_status(props, prop)


def generate_template_clicked(props, prop):
    global status_text
    try:
        from obs_template_generator import generate_template_from_scenes
        status_text = generate_template_from_scenes()
    except Exception as e:
        status_text = f"Template generation failed: {e}"
    print(status_text)
    return refresh_status(props, prop)


def script_properties():
    props = obs.obs_properties_create()

    obs.obs_properties_add_text(props, "status_info", status_text, obs.OBS_TEXT_INFO)

    db_list = obs.obs_properties_add_list(props, "db_format", "Data Source", obs.OBS_COMBO_TYPE_LIST, obs.OBS_COMBO_FORMAT_STRING)
    obs.obs_property_list_add_string(db_list, "Excel (.xlsm/.xlsx)", "excel")
    obs.obs_property_list_add_string(db_list, "CSV", "csv")
    obs.obs_property_list_add_string(db_list, "Google Sheet", "gsheet")

    file_prop = obs.obs_properties_add_path(props, "excel_file", "File Path", obs.OBS_PATH_FILE, "*.xlsm;*.xlsx;*.csv", None)
    # editable combos: auto-filled from the file, still typeable (needed for Google Sheets)
    ws_prop = obs.obs_properties_add_list(props, "worksheet", "Sheet Name/Label", obs.OBS_COMBO_TYPE_EDITABLE, obs.OBS_COMBO_FORMAT_STRING)
    obs.obs_properties_add_text(props, "gsheet_url", "Google Sheet URL", obs.OBS_TEXT_DEFAULT)
    obs.obs_properties_add_path(props, "gsheet_creds_path", "Service Account JSON (optional, for private sheets)", obs.OBS_PATH_FILE, "*.json", None)
    for key, label in zip(COL_KEYS, ("Source Type Column", "Source Name Column", "Value Column")):
        obs.obs_properties_add_list(props, key, label, obs.OBS_COMBO_TYPE_EDITABLE, obs.OBS_COMBO_FORMAT_STRING)
    # repopulate the dropdowns whenever the source file, format, or sheet changes
    obs.obs_property_set_modified_callback(db_list, refresh_pickers)
    obs.obs_property_set_modified_callback(file_prop, refresh_pickers)
    obs.obs_property_set_modified_callback(ws_prop, refresh_pickers)

    reload_interval_list = obs.obs_properties_add_list(props, "reload_interval", "Reload Interval (seconds)", obs.OBS_COMBO_TYPE_LIST, obs.OBS_COMBO_FORMAT_INT)
    obs.obs_property_list_add_int(reload_interval_list, "3 seconds", 3)
    obs.obs_property_list_add_int(reload_interval_list, "5 seconds", 5)
    obs.obs_property_list_add_int(reload_interval_list, "10 seconds", 10)

    obs.obs_properties_add_button(props, "toggle_script", "Start/Stop Script", toggle_script)
    obs.obs_properties_add_button(props, "refresh_status", "Refresh Status", refresh_status)
    obs.obs_properties_add_button(props, "generate_template", "Generate Source Template", generate_template_clicked)

    refresh_pickers(props, None, None)  # initial fill from the loaded settings
    return props


def script_defaults(settings):
    obs.obs_data_set_default_int(settings, "reload_interval", 3)
    obs.obs_data_set_default_string(settings, "db_format", "excel")


def script_update(settings):
    global excel_file, worksheet, source_type_column, source_name_column, value_column
    global reload_interval, db_format, gsheet_url, gsheet_creds_path, last_file_modified, data_cache
    db_format = obs.obs_data_get_string(settings, "db_format")
    excel_file = obs.obs_data_get_string(settings, "excel_file")
    worksheet = obs.obs_data_get_string(settings, "worksheet")
    gsheet_url = obs.obs_data_get_string(settings, "gsheet_url")
    gsheet_creds_path = obs.obs_data_get_string(settings, "gsheet_creds_path")
    source_type_column = obs.obs_data_get_string(settings, "source_type_column")
    source_name_column = obs.obs_data_get_string(settings, "source_name_column")
    value_column = obs.obs_data_get_string(settings, "value_column")
    reload_interval = obs.obs_data_get_int(settings, "reload_interval") or 3
    last_file_modified = 0  # settings changed: force a re-read next cycle
    data_cache = None


def script_load(settings):
    if deps_missing:
        install_deps(deps_missing)


def script_unload():
    global is_running
    is_running = False
    stop_event.set()
